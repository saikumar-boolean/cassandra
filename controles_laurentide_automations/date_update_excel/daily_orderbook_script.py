import pandas as pd
import os
import json
from dotenv import load_dotenv
import oracledb
import openpyxl
from datetime import datetime, timedelta
import traceback
import tempfile
import shutil

def main(context, utils):
    try:
        data_sheet_file_path = ""
        excel_file_name = ""
        temp_dir = None
        utils.set_context(context_obj=context)

        workbook = openpyxl.Workbook()

        POMatchesSheet = workbook.active
        POMatchesSheet.title = "PO Lines - Matches"
        POMatchesSheet["A1"] = "PO Order No"
        POMatchesSheet["B1"] = "Total Price"
        POMatchesSheet["C1"] = "Details"
        POMatchesSheet["D1"] = "Old Scheduled Date"
        POMatchesSheet["E1"] = "New Scheduled Date"

        COMatchesSheet = workbook.create_sheet(title="CO Lines - Matches")
        COMatchesSheet["A1"] = "CO Order No"
        COMatchesSheet["B1"] = "PO Order No"
        COMatchesSheet["C1"] = "CO Release No"
        COMatchesSheet["D1"] = "PO Total Price"
        COMatchesSheet["E1"] = "Details"
        COMatchesSheet["F1"] = "Old Promise Delivery Date"
        COMatchesSheet["G1"] = "New Promise Delivery Date"

        noMatches = workbook.create_sheet(title="Orderbook - No Match")
        noMatches["A1"] = "PO Order No"
        noMatches["B1"] = "Price"
        noMatches["C1"] = "Details"

        script_directory = os.path.dirname(os.path.abspath(__file__))
        config_file_path = os.path.join(script_directory, 'config.json')
        env_file_path = os.path.join(script_directory, '.env')
        temp_dir = tempfile.mkdtemp()
        
        utils.print_modified("env_file_path: " + env_file_path)

        load_dotenv(env_file_path)

        with open(config_file_path, 'r') as config_file:
            config = json.load(config_file)

        TENANT_ID = os.getenv("TENANT_ID")
        CLIENT_ID = os.getenv("CLIENT_ID")
        CLIENT_SECRET = os.getenv("CLIENT_SECRET")
        SENDER_EMAIL = os.getenv("SENDER_EMAIL")

        outlook_token = utils.get_graph_token(tenant_id=TENANT_ID, client_id=CLIENT_ID, client_secret=CLIENT_SECRET)
        
        current_email_timestamp = config["ORDERBOOK"]["current_email_timestamp"]
                
        data_sheet_file_path, current_email_timestamp = utils.get_latest_attachment_from_email_graph(token=outlook_token, mailbox_user_principal_name=os.getenv("PROJECT_EMAIL"), 
                                                                                                     outlook_folder="DAILY ORDERBOOK PSS", attachment_keyword="Daily Orderbook", 
                                                                                                     file_type="xlsm", email_timestamp=current_email_timestamp, 
                                                                                                     script_location=temp_dir)
        
        debug = config["ORDERBOOK"]["debug"]
        
        if(debug): recipients = os.getenv("DAILY_ORDERBOOK_EMAIL_LIST_TEST")
        else: recipients = os.getenv("DAILY_ORDERBOOK_EMAIL_LIST_PROD")
        
        if(data_sheet_file_path == "" and current_email_timestamp == ""):
            utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=SENDER_EMAIL, recipients=recipients, title="Daily Orderbook Automation", body_text=f"Most Recent file is already consolidated.")
            return
        
        data_sheet = pd.read_excel(data_sheet_file_path, usecols='B:AQ', sheet_name="Open Orders")
        
        oracle_username = os.getenv("PROD_USERNAME")
        oracle_password = os.getenv("PROD_PASSWORD")
        oracle_host = os.getenv("PROD_HOST")
        oracle_port = os.getenv("PROD_PORT")
        oracle_service_name = os.getenv("PROD_SERVICE_NAME")

        utils.print_modified("Oracle Info: " + oracle_username)

        if debug:
            utils.print_modified("Script in testing mode as IFS changes not commited.")

        connection = oracledb.connect(user=oracle_username, password=oracle_password, host=oracle_host, port=oracle_port, service_name=oracle_service_name)
        
        unique_order_nums = []
        data_sheet_rows = data_sheet.iloc

        for row in data_sheet_rows:
            if(' - ' in row['Repid']): row['Repid'] = row['Repid'].split(' - ')[1]
            elif('- ' in row['Repid']): row['Repid'] = row['Repid'].split('- ')[1]
            elif(' -' in row['Repid']): row['Repid'] = row['Repid'].split(' -')[1]
            elif('-' in row['Repid']): row['Repid'] = row['Repid'].split('-')[1]
            elif('--' in row['Repid']): row['Repid'] = row['Repid'].split('-')[1]
            else:
                utils.print_modified(f"Repid Value {row['Repid']} does not follow standard format.")
                continue
            
            if str(row['Repid']) + "-" + str(round(row['Unit Selling Price'] * abs(row["Ordered Quantity"]) , 2)) not in unique_order_nums: 
                unique_order_nums.append(str(row['Repid']) + "-" + str(round(row['Unit Selling Price'] * abs(row["Ordered Quantity"]) , 2)))

        matching_num_of_lines_counter = 0
        not_matching_num_of_lines_counter = 0
        total_counter = 0

        full_total_computed = []

        for unique_value in unique_order_nums:
            
            cursor = connection.cursor()
            split_value = str(unique_value).split("-")

            filtered_data_sheet = []
            for row in data_sheet_rows:
                if split_value[0] in row['Repid'] and split_value[1] == str(round(float(row["Unit Selling Price"]) * row["Ordered Quantity"], 2)): 
                    filtered_data_sheet.append(row)
            
            sql = f"SELECT (FBUY_UNIT_PRICE * BUY_QTY_DUE) as discounted_price, PCFT.CF$_SCHEDULED_DATE, POLT.DEMAND_ORDER_NO, POLT.DEMAND_RELEASE, POLT.* FROM PURCHASE_ORDER_LINE_TAB POLT inner join PURCHASE_ORDER_LINE_PART_CFT PCFT ON POLT.ROWKEY = PCFT.ROWKEY WHERE ROWSTATE != 'Closed' AND ORDER_NO = '{split_value[0]}' AND (FBUY_UNIT_PRICE * BUY_QTY_DUE) = {float(split_value[1])} ORDER BY PCFT.CF$_SCHEDULED_DATE"
            
            cursor.execute(sql)
            result = cursor.fetchall()
            is_not_present = all(sub_array[0] != split_value[0] for sub_array in full_total_computed)

            if (len(result) == len(filtered_data_sheet) and len(filtered_data_sheet) != 0 and [split_value[0], split_value[1] not in full_total_computed]): 
                utils.print_modified(f"Matching number of Excel rows and IFS rows at {len(result)} for {unique_value}")
                result = sorted(result, key=lambda x: (x[1] is None, x[1]))
                filtered_data_sheet = sorted(filtered_data_sheet, key=lambda y: y["Promise Date"])
                for i in range(0, len(result)):
                    promise_date = pd.to_datetime(filtered_data_sheet[i]['Promise Date'])
                    result_date = pd.to_datetime(result[i][1])
                    if(not pd.isna(promise_date) and not pd.isna(result_date) and promise_date.date() == result_date.date()):
                        utils.print_modified(f"Same Date: {result[i][1]}")
                        POMatchesSheet.append([split_value[0], float(split_value[1]), "Same Dates in IFS and Orderbook", result[i][1].strftime('%d/%m/%Y'), result[i][1].strftime('%d/%m/%Y')])
                        if(result[i][2] != None and result[i][3] != None):
                            sql = f"SELECT COLT.PROMISED_DELIVERY_DATE, COLT.* FROM CUSTOMER_ORDER_LINE_TAB COLT WHERE COLT.ORDER_NO = '{result[i][2]}' and COLT.REL_NO = {result[i][3]}"
                            
                            cursor.execute(sql)
                            result_CO = cursor.fetchall()

                            if(len(result_CO) == 1):
                                new_date = pd.to_datetime(result[i][1], format="%d/%m/%Y") + timedelta(days=14)
                                if(result_CO[0][0] == new_date):
                                    details = "Promised Date already changed."
                                else:
                                    details = "Promise Date changed by the script."
                                COMatchesSheet.append([result_CO[0][1], split_value[0], result[i][3], float(split_value[1]), details, result_CO[0][0].strftime('%d/%m/%Y'), new_date.strftime('%d/%m/%Y')])
                                cursor.execute(f"UPDATE CUSTOMER_ORDER_LINE_TAB SET PROMISED_DELIVERY_DATE = TO_DATE('{new_date.strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ORDER_NO = '{result[i][2]}' and REL_NO = {int(result[i][3])}")

                    elif (result[i][1] == None and (pd.isna(filtered_data_sheet[i]['Promise Date']) or filtered_data_sheet[i]['Promise Date'] is pd.NaT)):
                        utils.print_modified("Dates in IFS and Excel are empty")
                        POMatchesSheet.append([split_value[0], float(split_value[1]), "Empty Dates in IFS and Orderbook", "N/A", "N/A"])
                    else:
                        utils.print_modified("Different Date")
                        utils.print_modified(f"Excel Date: {filtered_data_sheet[i]['Promise Date']}")
                        utils.print_modified(f"IFS Date: {result[i][1]}")
                        ifsDate = "Empty" if result[i][1] == None else result[i][1].strftime('%d/%m/%Y')
                        newDate = "Empty" if pd.isna(filtered_data_sheet[i]['Promise Date']) else pd.to_datetime(promise_date).strftime("%d/%m/%Y")
                        
                        if(newDate == "Empty"):
                            details = "Different Dates in IFS and Orderbook - New Date not added since it is empty."
                        else:
                            details = "Different Dates in IFS and Orderbook - New Date Added."
                        POMatchesSheet.append([split_value[0], float(split_value[1]), details, ifsDate, newDate])

                        if(newDate != "Empty" and result[i][2] != None and result[i][2] != None):
                            sql = f"SELECT COLT.PROMISED_DELIVERY_DATE, COLT.* FROM CUSTOMER_ORDER_LINE_TAB COLT WHERE COLT.ORDER_NO = '{result[i][2]}' and COLT.REL_NO = {result[i][3]}"
                            cursor.execute(f"UPDATE PURCHASE_ORDER_LINE_PART_CFT SET CF$_SCHEDULED_DATE = TO_DATE('{newDate}', 'DD/MM/YYYY') WHERE ROWKEY = '{result[0][149]}'")

                            cursor.execute(sql)

                            result_CO = cursor.fetchall()

                            if(len(result_CO) == 1):
                                new_date = pd.to_datetime(newDate, format="%d/%m/%Y") + timedelta(days=14)

                                if(result_CO[0][0] == new_date):
                                    details = "Promised Date already changed."
                                else:
                                    details = "Promise Date changed by the script."
                                COMatchesSheet.append([result_CO[0][1], split_value[0], result[i][3], float(split_value[1]), details, result_CO[0][0].strftime('%d/%m/%Y'), new_date.strftime('%d/%m/%Y')])
                                cursor.execute(f"UPDATE CUSTOMER_ORDER_LINE_TAB SET PROMISED_DELIVERY_DATE = TO_DATE('{new_date.strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ORDER_NO = '{result[i][2]}' and REL_NO = {int(result[i][3])}")

                matching_num_of_lines_counter = matching_num_of_lines_counter + len(filtered_data_sheet)
                full_total_computed.append([split_value[0], split_value[1]])
            elif (is_not_present): 
                excel_sum = 0
                sql_sum = 0
                same_dates = True
                refiltered_data_sheet = []
                for excel_row in data_sheet_rows:
                    if split_value[0] in excel_row['Repid']: 
                        excel_sum = round(excel_sum + float(excel_row["Unit Selling Price"]) * row["Ordered Quantity"], 2)
                        refiltered_data_sheet.append(excel_row)
                
                sql = f"SELECT Sum(FBUY_UNIT_PRICE * BUY_QTY_DUE) FROM PURCHASE_ORDER_LINE_TAB POLT WHERE ROWSTATE != 'Closed' AND ORDER_NO = '{split_value[0]}'"
                
                cursor.execute(sql)
                sql_sum = cursor.fetchall()

                if (sql_sum[0][0] != None):
                    sql_sum = (str(sql_sum[0][0]))
                    sql_sum = float(sql_sum)

                    for date in refiltered_data_sheet: 
                        if (str(date["Schedule Ship Date"]) != "NaT" and refiltered_data_sheet[0]["Schedule Ship Date"] != date["Schedule Ship Date"]): 
                            same_dates = False

                    if (sql_sum == excel_sum and same_dates):
                        
                        utils.print_modified(f"Same Total Price: {sql_sum == excel_sum} - Same Dates: {same_dates} for Order No: {split_value[0]}")
                        matching_num_of_lines_counter = matching_num_of_lines_counter + len(refiltered_data_sheet)
                        full_total_computed.append([split_value[0], split_value[1]])
                        for row in data_sheet_rows:
                            if split_value[0] in row['Repid']:
                                newDate = (
                                    "Empty"
                                    if pd.isna(row['Promise Date'])
                                    else pd.to_datetime(row['Promise Date']).strftime("%d/%m/%Y")
                                )
                                POMatchesSheet.append([split_value[0], row["Unit Selling Price"] * row["Ordered Quantity"], f"Total from all lines with order no: {split_value[0]} - {split_value[1]}", "N/A", newDate])
                                if(len(result) == 0):
                                    sql = f"SELECT (POLT.FBUY_UNIT_PRICE * POLT.BUY_QTY_DUE) as discounted_price, PCFT.CF$_SCHEDULED_DATE, POLT.DEMAND_ORDER_NO, POLT.DEMAND_RELEASE, POLT.* FROM PURCHASE_ORDER_LINE_TAB POLT right join PURCHASE_ORDER_LINE_PART_CFT PCFT ON POLT.ROWKEY = PCFT.ROWKEY WHERE ROWSTATE != 'Closed' AND ORDER_NO = '{split_value[0]}' ORDER BY PCFT.CF$_SCHEDULED_DATE"
            
                                    cursor.execute(sql)
                                    result = cursor.fetchall()

                                if(newDate != "Empty" and len(result) != 0 and result[0][3] is not None and result[0][2] is not None):
                                    
                                    cursor.execute(f"UPDATE PURCHASE_ORDER_LINE_PART_CFT SET CF$_SCHEDULED_DATE = TO_DATE('{newDate}', 'DD/MM/YYYY') WHERE ROWKEY = '{result[0][149]}'")

                                    sql = f"SELECT COLT.PROMISED_DELIVERY_DATE, COLT.* FROM CUSTOMER_ORDER_LINE_TAB COLT WHERE COLT.ORDER_NO = '{result[0][2]}' and COLT.REL_NO = {int(result[0][3])}"
                                    
                                    cursor.execute(sql)
                                    result_CO = cursor.fetchall()
                                    
                                    if(len(result_CO) == 1 and result[0][1] != None):
                                        new_date = pd.to_datetime(newDate, format="%d/%m/%Y") + timedelta(days=14)
                                        if(result_CO[0][0] == new_date):
                                            details = "Promised Date already changed."
                                        else:
                                            details = "Promise Date changed by the script."
                                        COMatchesSheet.append([result_CO[0][1], split_value[0], result[0][3], float(split_value[1]), details, result_CO[0][0].strftime('%d/%m/%Y'), new_date.strftime('%d/%m/%Y')])
                                        cursor.execute(f"UPDATE CUSTOMER_ORDER_LINE_TAB SET PROMISED_DELIVERY_DATE = TO_DATE('{new_date.strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ORDER_NO = '{result[0][2]}' and REL_NO = {int(result[0][3])}")
                    else:
                        utils.print_modified(f"Same Total Price: {sql_sum == excel_sum} - Same Dates: {same_dates} for Order No: {split_value[0]}")
                        not_matching_num_of_lines_counter = not_matching_num_of_lines_counter + len(refiltered_data_sheet)
                        full_total_computed.append([split_value[0], split_value[1]])
                        for row in data_sheet_rows:
                            if split_value[0] in row['Repid']:
                                noMatches.append([split_value[0], row["Unit Selling Price"] * row["Ordered Quantity"], f"Did all rows in the order book have the same date: {same_dates} || Were the totals the same in IFS and order book: {sql_sum == excel_sum}"])
                        
                else:
                    utils.print_modified(f"No total amount value for {unique_value}")
                    not_matching_num_of_lines_counter = not_matching_num_of_lines_counter + len(refiltered_data_sheet)
                    full_total_computed.append([split_value[0], split_value[1]])
                    for row in data_sheet_rows:
                        if split_value[0] in row['Repid']:
                                noMatches.append([split_value[0], row["Unit Selling Price"] * row["Ordered Quantity"], f"Could not match using price value for all rows in order book."])

            total_counter += len(filtered_data_sheet)

        utils.print_modified(f"Number of lines matched: {matching_num_of_lines_counter}")
        utils.print_modified(f"Number of lines not matched: {not_matching_num_of_lines_counter}")
        utils.print_modified(f"Percentage match: %{matching_num_of_lines_counter / (matching_num_of_lines_counter + not_matching_num_of_lines_counter) * 100}")
        
        if (not debug):
            connection.commit()    
            
        connection.close()
        utils.update_json_value(config_file_path, ["ORDERBOOK", "current_email_timestamp"], current_email_timestamp)

        timestamp = datetime.now().strftime('%Y-%m-%d_%Hh-%Mm-%Ss')

        excel_file_name = f"results_orderbook_{timestamp}.xlsx"
        excel_file_path = os.path.join(temp_dir, excel_file_name)

        unique_combinations = set()
        rows_to_delete = []

        for row in range(2, COMatchesSheet.max_row + 1):
            co_order_no = COMatchesSheet[f"A{row}"].value
            co_release_no = COMatchesSheet[f"C{row}"].value
            combo = (co_order_no, co_release_no)

            # Check if combination already exists
            if combo in unique_combinations:
                rows_to_delete.append(row)  # Mark row for deletion
            else:
                unique_combinations.add(combo)

        # Delete duplicate rows (in reverse to prevent row shifting issues)
        for row in sorted(rows_to_delete, reverse=True):
            COMatchesSheet.delete_rows(row)

        utils.adjust_column_width(POMatchesSheet, workbook, excel_file_path)
        utils.adjust_column_width(COMatchesSheet, workbook, excel_file_path)
        utils.adjust_column_width(noMatches, workbook, excel_file_path)

        workbook.save(excel_file_path)
        workbook.close()

        utils.send_outlook_email_graph(token=outlook_token, attachement_name=excel_file_name, attachment_directory=temp_dir, mailbox_user_principal_name=SENDER_EMAIL,
                                       recipients=recipients, title="Daily Orderbook Automation Results", body_text="See attached the excel file indicating the changes in IFS.")
    except Exception as e:
        traceback.print_exc()
        utils.print_modified(str(e))
        env_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
        load_dotenv(env_file_path)
        utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=os.getenv("SENDER_EMAIL"), recipients=os.getenv("ADMIN_EMAIL"), title="Error! Daily Orderbook Automation", body_text=f"Error:\n\n{str(e)}\n\n{traceback.format_exc()}")
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            utils.print_modified("Temp dir deleted")
        utils.print_modified("Script completed")

if __name__ == "__main__":
    import utils as utils
    main(None, utils)