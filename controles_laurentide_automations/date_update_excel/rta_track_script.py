import pandas as pd
import os
import json
from dotenv import load_dotenv
import oracledb
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta, timezone
import traceback
from pathlib import Path
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
import re
import requests

def main(context, utils) -> tuple[bool, str]:
    try:
        connection = None
        now = datetime.now()
        #start_time_rta = f"{now.month}/{now.day}/{now.year}  {now.strftime('%I:%M:%S %p')}"
        utils.set_context(context)
        utils.print_modified("test")

        script_directory = os.path.dirname(os.path.abspath(__file__))
        config_file_path = os.path.join(script_directory, 'config.json')
        env_file_path = os.path.join(script_directory, '.env')
        
        utils.print_modified("env_file_path: " + env_file_path)

        load_dotenv(env_file_path)

        TENANT_ID = os.getenv("TENANT_ID")
        CLIENT_ID = os.getenv("CLIENT_ID")
        CLIENT_SECRET = os.getenv("CLIENT_SECRET")
        SENDER_EMAIL = os.getenv("RTA_SENDER_EMAIL")

        DISTANCE_PERCENTAGE = float(os.getenv("RTA_PERCENTAGE_LEVENSHTEIN_DISTANCE"))

        outlook_token = utils.get_graph_token(tenant_id=TENANT_ID, client_id=CLIENT_ID, client_secret=CLIENT_SECRET)

        excel_path = Path(
            os.getenv("RTA_TRACK_FILE_LOCATION")
        )

        with open(config_file_path, 'r') as config_file:
            config = json.load(config_file)

        debug = config["RTA"]["debug"]
        last_run_time = config["RTA"]["current_email_timestamp"]

        if(debug): recipients = os.getenv("RTA_TRACK_EMAIL_LIST_TEST")
        else: recipients = os.getenv("RTA_TRACK_EMAIL_LIST_PROD")
                

        rta_track_file = pd.ExcelFile(excel_path)

        names_file_path = Path(
                os.getenv("RTA_TRACK_NAME_CHECK_FILE")
            )

        names_file = pd.read_excel(names_file_path)

        for sheet in rta_track_file.sheet_names:
            if sheet == "Output":
                df_excel = pd.read_excel(rta_track_file, sheet_name=sheet)
                utils.print_modified(f"Sheet: {sheet}, shape: {df_excel.shape}")
                utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=SENDER_EMAIL, recipients=recipients, title="RTA Invoice Load Queue Alert", body_text=f"Le dernier rapport Track a déjà été rapproché par la solution.\n\nVeuillez utiliser un nouveau fichier.\n\nThe newest Track report has already been reconciled by the solution.\n\nPlease use a new file.")
                return

        oracle_username = os.getenv("PROD_USERNAME")
        oracle_password = os.getenv("PROD_PASSWORD")
        oracle_host = os.getenv("PROD_HOST")
        oracle_port = os.getenv("PROD_PORT")
        oracle_service_name = os.getenv("PROD_SERVICE_NAME")

        if debug:
            utils.print_modified("Script in testing mode, changes in IFS not commited.")

        connection = oracledb.connect(user=oracle_username, password=oracle_password, host=oracle_host, port=oracle_port, service_name=oracle_service_name)
        cursor = connection.cursor()

        customer_number = 10462

        project_transaction_select_query = f"SELECT E.EMP_NO || ' - ' || E.EXTERNAL_DISPLAY_NAME as Employee, PT.ACTIVITY_NO || ' - ' || PA.DESCRIPTION Activity, PT.ACCOUNT_DATE, PT.INTERNAL_QUANTITY,  ROUND(PT.INTERNAL_PRICE, 2) AS INTERNAL_PRICE, PT.INTERNAL_AMOUNT, ROUND(PT.SALES_PRICE, 2) AS SALES_PRICE, ROUND(PT.SALES_QUANTITY, 2) AS SALES_QUANTITY, PT.INVOICABILITY,  PT.INVOICE_ID, PT.PROJECT_TRANSACTION_SEQ, PT.PROJECT_ID, PT.SUB_PROJECT_ID,  PT.ACTIVITY_SEQ, PT.SHORT_NAME, PT.C_PO_NO, PI.CLIENT_STATE AS INVOICE_STATUS  FROM {oracle_username}.PROJECT_TRANSACTION PT  LEFT JOIN {oracle_username}.PROJECT_INVOICE PI ON PT.INVOICE_ID = PI.INVOICE_ID  RIGHT JOIN {oracle_username}.COMPANY_PERSON_PUB E ON PT.EMP_NO = E.EMP_NO LEFT JOIN {oracle_username}.PROJECT_ACTIVITY PA ON PT.ACTIVITY_SEQ = PA.ACTIVITY_SEQ  WHERE PT.Customer = '{customer_number}'"

        df_ifs = pd.read_sql(project_transaction_select_query, connection)
        df_ifs["Name Score"] = None
        df_ifs["Rate Score"] = None
        df_ifs["Hours Difference"] = None
        df_ifs["Update Hours"] = None
        df_ifs["Result"] = None

        df_excel = pd.read_excel(excel_path)

        for col in df_excel.columns:
            utils.print_modified(f"'{col}'")
        
        df_excel.columns = df_excel.columns.str.rstrip()
        df_excel.columns = df_excel.columns.str.replace(".", "", regex=False)

        df_excel = df_excel[["date de feuille de temps", "nom", "prénom", "État de l’attrib", 
                "taux 1", "heures à taux 1", "taux 2", "heures à taux 2", "taux 3",
                "heures à taux 3", "date de la dernière modification", "Total général ($)", 
                "dernière modification par", "référence de la feuille de temps", 
                "id de projet virtuel"]]

        df_excel["Reference"] = df_excel.apply(
            lambda row: f"{row['date de feuille de temps']}-{row['id de projet virtuel']}-{row['Total général ($)']}-{row['nom']}",
            axis=1
        )

        df_excel["ExcelRow"] = (df_excel.index + 2).astype(str)

        df_excel = df_excel.reset_index(drop=True)
        
        for excel_itt, row in df_excel.iterrows():
            df_excel.loc[excel_itt, "is_field_empty"] = False
            df_excel.loc[excel_itt, "empty_field"] = ''
            df_excel.loc[excel_itt, "is_non_billable"] = False
            df_excel.loc[excel_itt, "Filter Out"] = True
            df_excel.loc[excel_itt, "Exception"] = False
            df_excel.loc[excel_itt, "Case ID"] = str(row["date de feuille de temps"]) + "-" + str(row["id de projet virtuel"]) + '-' + utils.format_to_money(row["Total général ($)"]) + " -" + str(row["nom"])
            df_excel.loc[excel_itt, "Hours Difference"] = 0
            df_excel.loc[excel_itt, "Process"] = "RTA Invoicing"
            now = datetime.now()
            formatted = f"{now.month}/{now.day}/{now.year}  {now.strftime('%I:%M:%S %p')}"

            df_excel.loc[excel_itt, "Case Creation Time"] = formatted
            df_excel.loc[excel_itt, "Case Start Time"] = formatted
            df_excel.loc[excel_itt, "Case SLA Time"] = ''
            df_excel.loc[excel_itt, "Case Defer Time"] = ''
            df_excel.loc[excel_itt, "Work Time"] = ''
            df_excel.loc[excel_itt, "Case End Time"] = ''

            #d = datetime.strptime(str(row["date de la dernière modification"]), "%Y%m%d").date()
            d = datetime.strptime(str(row["date de feuille de temps"]), "%Y%m%d").date()
            
            if (date.today() - d) >= timedelta(days=21) and str(row["État de l’attrib"]) != "Authorized":
                utils.print_modified(f"Business Exception")
                utils.print_modified("Non-Authorized transaction is more than 21 days old.")
                df_excel.loc[excel_itt, "Result"] = "Non-Authorized transaction is more than 21 days old."
                df_excel.loc[excel_itt, "Case Result"] = "Business Exception."
                df_excel.loc[excel_itt, "Exception Reason"] = "Non-Authorized transaction is more than 21 days old."
                
                df_excel.loc[excel_itt, "Filter Out"] = False
                df_excel.loc[excel_itt, "Exception"] = True
                now = datetime.now()
                formatted = f"{now.month}/{now.day}/{now.year}  {now.strftime('%I:%M:%S %p')}"

                df_excel.loc[excel_itt, "Case End Time"] = formatted
                continue

            empty_fields = row[row.isna() | (row == '')]

            if empty_fields.size != 0 and str(row["État de l’attrib"]) == "Authorized":
                empty_columns = empty_fields.index.tolist()
                utils.print_modified("Business Exception")
                utils.print_modified(f"Missing field in Transaction Item: {empty_columns}")
                df_excel.loc[excel_itt, "empty_field"] = "" + ', '.join(empty_columns)
                df_excel.loc[excel_itt, "Result"] = f"Missing field in Transaction Item: {empty_columns}"
                df_excel.loc[excel_itt, "Case Result"] = "Business Exception."
                df_excel.loc[excel_itt, "Exception Reason"] = f"Missing field in Transaction Item: {empty_columns}"


                df_excel.loc[excel_itt, "Filter Out"] = False
                df_excel.loc[excel_itt, "Exception"] = True
                now = datetime.now()
                formatted = f"{now.month}/{now.day}/{now.year}  {now.strftime('%I:%M:%S %p')}"

                df_excel.loc[excel_itt, "Case End Time"] = formatted
                continue

            taux_1 = utils.convert_from_money(row["taux 1"])
            taux_2 = utils.convert_from_money(row["taux 2"])
            taux_3 = utils.convert_from_money(row["taux 3"])

            heurs_1 = utils.convert_from_money(row["heures à taux 1"])
            heurs_2 = utils.convert_from_money(row["heures à taux 2"])
            heurs_3 = utils.convert_from_money(row["heures à taux 3"])

            if taux_1 + taux_2 + taux_3 == 0 and heurs_1 + heurs_2 + heurs_3 != 0:
                df_excel.loc[excel_itt, "is_non_billable"] = True
                df_excel.loc[excel_itt, "PO Number"] = "Unavailable"
                df_excel.loc[excel_itt, "Result"] = "No Match: Non-billable transaction"
                df_excel.loc[excel_itt, "Nearest Match Employee"] = ""
                df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                df_excel.loc[excel_itt, "ShortName"] = ""
                df_excel.loc[excel_itt, "Hours Difference"] = 0
                df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                df_excel.loc[excel_itt, "Filter Out"] = False
                utils.print_modified("Transaction is non-billable.  Skipping reconciliation.")
                continue

            df_excel.loc[excel_itt, "référence de la feuille de temps modifié"] = (
                str(row["référence de la feuille de temps"]).replace(",", "").replace(".", "")
            )

            if str(row["État de l’attrib"]) != "Authorized":
                df_excel.loc[excel_itt, "Match"] = False
                df_excel.loc[excel_itt, "PO Number"] = "Unavailable"
                df_excel.loc[excel_itt, "Result"] = "No Match: Status not Authorized"
                df_excel.loc[excel_itt, "Nearest Match Employee"] = ""
                df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                df_excel.loc[excel_itt, "ShortName"] = ""
                df_excel.loc[excel_itt, "Hours Difference"] = 0
                df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                df_excel.loc[excel_itt, "Filter Out"] = False
                utils.print_modified("No Match: Status not Authorized")
                continue
            else:
                utils.print_modified("Check PO Num")
                df_ifs_po_filtered = df_ifs[df_ifs["C_PO_NO"].astype(str) == str(row["id de projet virtuel"])].copy()
                df_ifs_po_filtered.reset_index(drop=True, inplace=True)

                if len(df_ifs_po_filtered) == 0:
                    df_excel.loc[excel_itt, "Match"] = False
                    df_excel.loc[excel_itt, "PO Number"] = "Unavailable"
                    df_excel.loc[excel_itt, "Result"] = "No Match: PO"
                    df_excel.loc[excel_itt, "Nearest Match Employee"] = ""
                    df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                    df_excel.loc[excel_itt, "ShortName"] = ""
                    df_excel.loc[excel_itt, "Hours Difference"] = 0
                    df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                    df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                    df_excel.loc[excel_itt, "Filter Out"] = False
                    utils.print_modified("No Match: PO")
                    continue

                row_date_str = str(row["date de feuille de temps"])
                row_date = datetime.strptime(row_date_str, "%Y%m%d")

                df_ifs_po_filtered["ACCOUNT_DATE"] = pd.to_datetime(
                    df_ifs_po_filtered["ACCOUNT_DATE"],
                    errors="coerce"
                )
                df_ifs_date_filtered = df_ifs_po_filtered[
                    df_ifs_po_filtered["ACCOUNT_DATE"] == row_date
                ].copy()
                df_ifs_date_filtered.reset_index(drop=True, inplace=True)

                if len(df_ifs_date_filtered) == 0:
                    df_excel.loc[excel_itt, "Match"] = False
                    df_excel.loc[excel_itt, "PO Number"] = row["id de projet virtuel"]
                    df_excel.loc[excel_itt, "Result"] = "No Match: Account Date"
                    df_excel.loc[excel_itt, "Nearest Match Employee"] = ""
                    df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                    df_excel.loc[excel_itt, "ShortName"] = ""
                    df_excel.loc[excel_itt, "Hours Difference"] = 0
                    df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                    df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                    df_excel.loc[excel_itt, "Filter Out"] = False
                    utils.print_modified("No Match: Account Date")
                    continue

                if str(round(row["heures à taux 1"], 2)) == "0.0":
                    track_rate = Decimal(str(row["taux 2"]).replace("$", "")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                else:
                    track_rate = Decimal(str(row["taux 1"]).replace("$", "")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                
                df_ifs_rate_filtered = df_ifs_date_filtered[
                    df_ifs_date_filtered["SALES_PRICE"].apply(lambda x: Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) == track_rate)
                ].copy()
                df_ifs_rate_filtered.reset_index(drop=True, inplace=True)

                if len(df_ifs_rate_filtered) == 0:
                    df_excel.loc[excel_itt, "Match"] = False
                    df_excel.loc[excel_itt, "PO Number"] = row["id de projet virtuel"]
                    df_excel.loc[excel_itt, "Result"] = "No Match: Sales Price (rate)"
                    df_excel.loc[excel_itt, "Nearest Match Employee"] = ""
                    df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                    df_excel.loc[excel_itt, "ShortName"] = ""
                    df_excel.loc[excel_itt, "Hours Difference"] = 0
                    df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                    df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                    df_excel.loc[excel_itt, "Filter Out"] = False
                    utils.print_modified("No Match: Sales Price (rate)")
                    continue

                for _, ifs_row in df_ifs_rate_filtered.iterrows():
                    employee = ifs_row["EMPLOYEE"]
                    ifs_name = employee.split(" - ", 1)[1].strip()
                    track_name = str(row["prénom"]) + " " + str(row["nom"])

                    name_match = False
                    name_score = ""

                    for __, current_row in names_file.iterrows():
                        track_sheet_name = str(current_row["Prénom TRACK"]) + " " + str(current_row["Nom TRACK"])
                        ifs_sheet_name = str(current_row["Prénom IFS"]) + " " + str(current_row["Nom IFS"])

                        if track_sheet_name == track_name:
                            if ifs_name == ifs_sheet_name:
                                name_match = True
                                name_score = "Match"
                                break
                            else:
                                name_match = False

                    if not name_match:
                        name_match, name_score = utils.levenshtein_distance(
                            ifs_name, track_name, DISTANCE_PERCENTAGE
                        )

                    df_ifs_rate_filtered.loc[_, "Name Score"] = name_score
                    df_ifs_rate_filtered.loc[_, "Result"] = "Match" if name_match else "No Match: Name"
            
            df_ifs_name_filtered = df_ifs_rate_filtered[df_ifs_rate_filtered["Result"] == "Match"]
            df_ifs_name_filtered.reset_index(drop=True, inplace=True)

            if len(df_ifs_name_filtered) == 0:
                df_ifs_rate_filtered.sort_values(by="Name Score", ascending=False, inplace=True)
                msg = (
                    f"No Match: Employee Name - Track Name: "
                    f"{df_excel.loc[excel_itt, 'prénom']} "
                    f"{df_excel.loc[excel_itt, 'nom']}. "
                    f"Closest IFS Name: {df_ifs_rate_filtered.loc[0, 'EMPLOYEE']}"
                )
                df_excel.loc[excel_itt, "Match"] = False
                df_excel.loc[excel_itt, "PO Number"] = row["id de projet virtuel"]
                df_excel.loc[excel_itt, "Result"] = msg
                df_excel.loc[excel_itt, "Nearest Match Employee"] = df_ifs_rate_filtered.loc[0, "EMPLOYEE"]
                df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                df_excel.loc[excel_itt, "ShortName"] = ""
                df_excel.loc[excel_itt, "Hours Difference"] = 0
                df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                df_excel.loc[excel_itt, "Filter Out"] = False
                utils.print_modified(msg)
                continue

            elif len(df_ifs_name_filtered) > 1:
                df_excel.loc[excel_itt, "Match"] = False
                df_excel.loc[excel_itt, "PO Number"] = row["id de projet virtuel"]
                df_excel.loc[excel_itt, "Result"] = "Multiple matches found"
                df_excel.loc[excel_itt, "Nearest Match Employee"] = df_ifs_rate_filtered.loc[0, "EMPLOYEE"]
                df_excel.loc[excel_itt, "Nearest Transaction Activity"] = ""
                df_excel.loc[excel_itt, "ShortName"] = ""
                df_excel.loc[excel_itt, "Hours Difference"] = 0
                df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = ""
                df_excel.loc[excel_itt, "IFS Sales Amount"] = ""
                df_excel.loc[excel_itt, "Filter Out"] = False
                utils.print_modified("Multiple matches found")
                continue
            else:
                if str(round(row["heures à taux 1"], 2)) == "0.0":
                    #track_hours = Decimal(str(row["heures à taux 2"])).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
                    track_hours = utils.normalize_numeric(row["heures à taux 2"])
                else:
                    #track_hours = Decimal(str(row["heures à taux 1"])).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
                    track_hours = utils.normalize_numeric(row["heures à taux 1"])

                ifs_hours = utils.normalize_numeric(df_ifs_name_filtered.loc[0, "SALES_QUANTITY"])

                if track_hours == ifs_hours:
                    match_reason = "Match - hours were the same in IFS and Track"
                    were_hours_updates = False
                    utils.print_modified(match_reason)
                else:
                    sales_quantity_update_query = f'UPDATE PROJECT_TRANSACTION_TAB SET SALES_QUANTITY = {track_hours} WHERE PROJECT_TRANSACTION_SEQ = {int(df_ifs_name_filtered.loc[0, "PROJECT_TRANSACTION_SEQ"])}'
                    cursor.execute(sales_quantity_update_query)

                    precise_difference = (Decimal(str(ifs_hours)) - Decimal(str(track_hours))).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
                    ifs_track_hours_difference_update_query = f'UPDATE PROJECT_TRANSACTION_TAB SET IFS_TRACK_HOURS_DIFFERENCE = {precise_difference} WHERE PROJECT_TRANSACTION_SEQ = {int(df_ifs_name_filtered.loc[0, "PROJECT_TRANSACTION_SEQ"])}'
                    cursor.execute(ifs_track_hours_difference_update_query)

                    match_reason = (
                        "Match - Hours from IFS and Track were different and were modified to reflect the hours in Track"
                        + " - IFS hours before modifications: " + str(ifs_hours) + " - Track hours: " + str(track_hours)
                    )
                    were_hours_updates = True
                    utils.print_modified(match_reason)

                select_specific_project_transaction = f'SELECT INVOICE_COMMENTS, SALES_AMOUNT FROM PROJECT_TRANSACTION WHERE PROJECT_TRANSACTION_SEQ = {int(df_ifs_name_filtered.loc[0, "PROJECT_TRANSACTION_SEQ"])}'
                cursor.execute(select_specific_project_transaction)
                query_results = cursor.fetchall()
                
                invoice_comments = str(query_results[0][0])
                sales_amount_value = query_results[0][1] or 0

                invoice_comments_match = re.search(r"[0-9]{7}", invoice_comments) is not None

                if invoice_comments_match:
                    df_excel.loc[excel_itt, "Filter Out"] = True
                    df_excel.loc[excel_itt, "Invoice Comments Details"] = ""
                else:
                    df_excel.loc[excel_itt, "Filter Out"] = False
                    invoice_comments_update_query = f'UPDATE PROJECT_TRANSACTION_TAB SET INVOICE_COMMENTS = {int(float(row["référence de la feuille de temps"]))} WHERE PROJECT_TRANSACTION_SEQ = {int(df_ifs_name_filtered.loc[0, "PROJECT_TRANSACTION_SEQ"])}'
                    cursor.execute(invoice_comments_update_query)
                    if invoice_comments and invoice_comments.strip() and invoice_comments != 'None':
                       df_excel.loc[excel_itt, "Invoice Comments Details"] = f" - Invoice comment changed to " + str(int(float(row["référence de la feuille de temps"]))) + " - original value was: " + str(invoice_comments) 
                    else:
                        df_excel.loc[excel_itt, "Invoice Comments Details"] = f" - Invoice comment changed to " + str(int(float(row["référence de la feuille de temps"]))) + " - original value was: No Invoice comment"

                sales_amount = [Decimal(str(sales_amount_value))]
                hours_difference = f"{abs(ifs_hours - track_hours):,.2f}"

                if were_hours_updates:
                    df_excel.loc[excel_itt, "Filter Out"] = False

                df_excel.loc[excel_itt, "Match"] = True
                df_excel.loc[excel_itt, "PO Number"] = row["id de projet virtuel"]
                df_excel.loc[excel_itt, "Result"] = match_reason + df_excel.loc[excel_itt, "Invoice Comments Details"]
                df_excel.loc[excel_itt, "Nearest Match Employee"] = df_ifs_rate_filtered.loc[0, "EMPLOYEE"]
                df_excel.loc[excel_itt, "Nearest Transaction Activity"] = df_ifs_rate_filtered.loc[0, "ACTIVITY"]
                df_excel.loc[excel_itt, "ShortName"] = df_ifs_rate_filtered.loc[0, "SHORT_NAME"]
                df_excel.loc[excel_itt, "Hours Difference"] = hours_difference
                df_excel.loc[excel_itt, "ProjectTransactionSEQ"] = str(df_ifs_name_filtered.loc[0, "PROJECT_TRANSACTION_SEQ"])
                df_excel.loc[excel_itt, "IFS Sales Amount"] = f"${sales_amount[0]:,.2f}"

            if df_excel.loc[excel_itt, "Match"] is True:
                utils.print_modified("Match!")

        normal_cols = ["date de feuille de temps", "nom", "prénom", "État de l’attrib", "taux 1", "heures à taux 1", "taux 2", "heures à taux 2", "taux 3", "heures à taux 3", "date de la dernière modification", "dernière modification par", "référence de la feuille de temps", "id de projet virtuel", "Total général ($)"]
        result_col = "Result"
        
        df_excel["heures à taux 1"] = df_excel["heures à taux 1"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=False)
        )
        df_excel["heures à taux 2"] = df_excel["heures à taux 2"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=False)
        )
        df_excel["heures à taux 3"] = df_excel["heures à taux 3"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=False)
        )

        df_excel["Total général ($)"] = df_excel["Total général ($)"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=True)
        )
        df_excel["taux 1"] = df_excel["taux 1"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=True)
        )
        df_excel["taux 2"] = df_excel["taux 2"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=True)
        )
        df_excel["taux 3"] = df_excel["taux 3"].apply(
            lambda x: utils.format_to_money(x, dollar_sign=True)
        )

        df_excel["ProjectTransactionSEQ"] = df_excel["ProjectTransactionSEQ"].apply(
            lambda v: "" if v is None or v == "" else ("'" + str(v))
        )

        df_excel["PO Number"] = df_excel["PO Number"].apply(
            lambda v: "" if v is None or v == "" else ("'" + str(v))
        )

        try:
            wb = load_workbook(excel_path)
        except FileNotFoundError:
            wb = Workbook()

        if "Output" in wb.sheetnames:
            ws = wb["Output"]
            wb.remove(ws)

        ws = wb.create_sheet(title="Output")

        for col_idx, col_name in enumerate(normal_cols, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, val in enumerate(df_excel[col_name], start=2):
                ws.cell(row=row_idx, column=col_idx, value=val)

        special_col_excel_idx = 18
        ws.cell(row=1, column=special_col_excel_idx, value=result_col)
        for row_idx, val in enumerate(df_excel[result_col], start=2):
            ws.cell(row=row_idx, column=special_col_excel_idx, value=val)

        wb.save(excel_path)

        utils.print_modified(df_excel.head())

        df_excel["Total général ($)"] = (
            df_excel["Total général ($)"]
            .astype(str)
            .str.replace("$", "", regex=False)
            .str.strip()        
        )
        df_excel["taux 1"] = (
            df_excel["taux 1"]
            .astype(str)
            .str.replace("$", "", regex=False)
            .str.strip()        
        )
        df_excel["taux 2"] = (
            df_excel["taux 2"]
            .astype(str)
            .str.replace("$", "", regex=False)
            .str.strip()        
        )
        df_excel["taux 3"] = (
            df_excel["taux 3"]
            .astype(str)
            .str.replace("$", "", regex=False)
            .str.strip()        
        )
        
        SP_HOSTNAME = os.getenv("SP_HOSTNAME")
        SP_SITE_PATH = os.getenv("SP_SITE_PATH")
        TEMPLATE_FILENAME = os.getenv("RTA_TRACK_EMAIL_TEMPLATE_SHAREPOINT")

        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
        token_data = {
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "scope": "https://graph.microsoft.com/.default"
        }
        token_resp = requests.post(token_url, data=token_data)
        token_resp.raise_for_status()
        access_token = token_resp.json()["access_token"]

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        site_url = f"https://graph.microsoft.com/v1.0/sites/{SP_HOSTNAME}:{SP_SITE_PATH}"
        site_id = requests.get(site_url, headers=headers).json()["id"]

        search_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{TEMPLATE_FILENAME}')"
        search_resp = requests.get(search_url, headers=headers)
        search_resp.raise_for_status()
        items = search_resp.json()["value"]
        template_item = next(x for x in items if x["name"] == TEMPLATE_FILENAME)

        template_id = template_item["id"]
        parent_id = template_item["parentReference"]["id"]

        new_name = f"RTA Invoicing MI Report.xlsx"
        copy_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{template_id}/copy"
        copy_body = {
            "parentReference": {"id": parent_id},
            "name": new_name
        }

        copy_resp = requests.post(copy_url, headers=headers, json=copy_body)
        copy_resp.raise_for_status()

        monitor_url = copy_resp.headers.get("Location")
        if not monitor_url:
            raise RuntimeError("Missing monitor URL from copy response")

        file_id = utils.wait_for_copy_completion(monitor_url=monitor_url)

        df_filtered = df_excel[df_excel["Filter Out"] == False]
        df_success = df_filtered[df_filtered["Exception"] == False]
        df_exceptions = df_filtered[df_filtered["Exception"] == True]

        success_columns = [
            "Case ID","ProjectTransactionSEQ","Result","date de feuille de temps",
            "prénom","nom","PO Number","heures à taux 1","taux 1",
            "heures à taux 2","taux 2","heures à taux 3","taux 3","Total général ($)","IFS Sales Amount",
            "Hours Difference","État de l’attrib","référence de la feuille de temps",
            "dernière modification par","date de la dernière modification"
        ]

        exception_columns = [
            "Process","Case ID","Case Creation Time","Case SLA Time",
            "Case Start Time","Case End Time","Case Defer Time","Case Result",
            "Exception Reason","Work Time"
        ]
        now = datetime.now()
        end_time_rta = f"{now.month}/{now.day}/{now.year}  {now.strftime('%I:%M:%S %p')}"

        dt = datetime.fromisoformat(last_run_time)
        last_run_time_formatted = f"{dt.month}/{dt.day}/{dt.year}  {dt.strftime('%I:%M:%S %p')}"

        utils.write_excel_range(outlook_token, site_id, file_id, "Dashboard", "T3", last_run_time_formatted)
        utils.write_excel_range(outlook_token, site_id, file_id, "Dashboard", "T4", end_time_rta)
        
        utils.update_sheet("Successes", "A4", df_success, success_columns, site_id, file_id, headers)
        if len(df_exceptions) > 0:
            utils.update_sheet("Exceptions", "A4", df_exceptions, exception_columns, site_id, file_id, headers)
        
        utils.set_excel_column_widths(
            access_token=access_token,
            site_id=site_id,
            file_id=file_id,
            sheet_name="Successes",
            columns = [
                "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"
            ],
            widths = [
                101, 130, 101, 97, 120, 119, 99, 97, 91, 97,
                87, 97, 87, 91, 108, 112, 91, 117, 104, 103
            ]
        )
        utils.set_excel_column_widths(
            access_token=access_token,
            site_id=site_id,
            file_id=file_id,
            sheet_name="Exceptions",
            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
            widths = [120, 388, 197, 156, 167, 162, 170, 166, 438, 119]
        )
        
        utils.refresh_pivots_in_sheet(outlook_token, site_id, file_id, "Dashboard")
        
        body_html = os.getenv("RTA_EMAIL_BODY")
        signature_html = os.getenv("RTA_EMAIL_SIGNATURE")

        utils.send_outlook_email_with_sharepoint_attachment(
            token=outlook_token,
            mailbox_user_principal_name=SENDER_EMAIL,
            site_id=site_id,
            file_id=file_id,
            recipients=recipients,
            title="RTA Invoicing MI Report",
            body_html=body_html,
            signature_html=signature_html
        )
        
        if (not debug):
            utils.print_modified("Changes committed in IFS")
            connection.commit()
            utils.update_json_value(config_file_path, ["RTA", "current_email_timestamp"], datetime.now(timezone.utc).isoformat(timespec="seconds"))

        utils.print_modified("RTA Track Script Finished")
        
    except Exception as e:
        traceback.print_exc()
        utils.print_modified(str(e))
        env_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
        load_dotenv(env_file_path)
        utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=os.getenv("RTA_SENDER_EMAIL"), recipients=os.getenv("ADMIN_EMAIL"), title="Error! Larentide RTA Track Automation", body_text=f"Error:\n\n{str(e)}\n\n{traceback.format_exc()}")
    finally:
        if connection is not None:
            connection.close() 

if __name__ == "__main__":
    import utils as utils
    main(None, utils)