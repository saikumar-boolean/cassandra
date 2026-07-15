import pandas as pd
import os
import json
from dotenv import load_dotenv
import oracledb
import openpyxl
from datetime import datetime, timedelta
import traceback
import tempfile
import shutil

def main(context, utils):
    try:
        data_sheet_file_path = ""
        excel_file_name = ""
        temp_dir = None
        utils.set_context(context_obj=context)     
        
        workbook = openpyxl.Workbook()
        
        POMatchesSheet = workbook.active
        POMatchesSheet.title = "PO Lines - Matches"
        POMatchesSheet["A1"] = "PO Order No"
        POMatchesSheet["B1"] = "Part No"
        POMatchesSheet["C1"] = "Details"
        POMatchesSheet["D1"] = "Old Scheduled Date"
        POMatchesSheet["E1"] = "New Scheduled Date"
        POMatchesSheet["F1"] = "Number of Lines in IFS Changed"
        POMatchesSheet["G1"] = "Number of Lines in Excel sheet with this OrderNo-PartNo Combination"

        COMatchesSheet = workbook.create_sheet(title="CO Lines - Matches")
        COMatchesSheet["A1"] = "CO Order No"
        COMatchesSheet["B1"] = "PO Order No"
        COMatchesSheet["C1"] = "CO Release No"
        COMatchesSheet["D1"] = "CO Line No"
        COMatchesSheet["E1"] = "PO Part No"
        COMatchesSheet["F1"] = "Details"
        COMatchesSheet["G1"] = "Old Promise Delivery Date"
        COMatchesSheet["H1"] = "New Promise Delivery Date"
        COMatchesSheet["I1"] = "Number of Lines in IFS Changed"


        noMatches = workbook.create_sheet(title="Orderbook - No Match")
        noMatches["A1"] = "PO Order No"
        noMatches["B1"] = "Excel Part No"
        noMatches["C1"] = "Details"

        script_directory = os.path.dirname(os.path.abspath(__file__))
        config_file_path = os.path.join(script_directory, 'config.json')
        env_file_path = os.path.join(script_directory, '.env')
        temp_dir = tempfile.mkdtemp()
        
        utils.print_modified("env_file_path: " + env_file_path)

        load_dotenv(env_file_path)

        with open(config_file_path, 'r') as config_file:
            config = json.load(config_file)

        TENANT_ID = os.getenv("TENANT_ID")
        CLIENT_ID = os.getenv("CLIENT_ID")
        CLIENT_SECRET = os.getenv("CLIENT_SECRET")
        SENDER_EMAIL = os.getenv("SENDER_EMAIL")

        outlook_token = utils.get_graph_token(tenant_id=TENANT_ID, client_id=CLIENT_ID, client_secret=CLIENT_SECRET)

        current_email_timestamp = config["PARTNER"]["current_email_timestamp"]
        data_sheet_file_path, current_email_timestamp = utils.get_latest_attachment_from_email_graph(token=outlook_token, mailbox_user_principal_name=os.getenv("PROJECT_EMAIL"), 
                                                                                                     outlook_folder="Clint Emerson Rosemount report", attachment_keyword="Order_Status_Report", 
                                                                                                     file_type="xlsx", email_timestamp=current_email_timestamp, 
                                                                                                     script_location=temp_dir)
        
        utils.print_modified(f"data_sheet_file_path: {data_sheet_file_path}")
        utils.print_modified(f"current_email_timestamp: {current_email_timestamp}")

        debug = config["PARTNER"]["debug"]

        if(debug): recipients = os.getenv("IMPACT_PARTNER_EMAIL_LIST_TEST")
        else: recipients = os.getenv("IMPACT_PARTNER_EMAIL_LIST_PROD")
            
        if(data_sheet_file_path == "" and current_email_timestamp == ""):
            utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=SENDER_EMAIL, recipients=recipients, title="Impact Partner Excel Sheet Automation", body_text=f"Most Recent file is already consolidated.")
            return

        utils.update_json_value(config_file_path, ["PARTNER", "current_email_timestamp"], current_email_timestamp)

        data_sheet = pd.read_excel(data_sheet_file_path, usecols='A:AU', sheet_name="Canada Impact Partenrs Order St")   


        oracle_username = os.getenv("PROD_USERNAME")
        oracle_password = os.getenv("PROD_PASSWORD")
        oracle_host = os.getenv("PROD_HOST")
        oracle_port = os.getenv("PROD_PORT")
        oracle_service_name = os.getenv("PROD_SERVICE_NAME")

        if debug:
            utils.print_modified("Script in testing mode, changes in IFS not commited.")

        connection = oracledb.connect(user=oracle_username, password=oracle_password, host=oracle_host, port=oracle_port, service_name=oracle_service_name)
        
        unique_order_nums = []
        data_sheet_rows = data_sheet.iloc
        
        for row in data_sheet_rows:
            if(str(row['PO']) + "-" + str(row['MODEL']) in unique_order_nums):
                utils.print_modified(f"Order No {row['PO']} with Model {row['MODEL']} is a duplicate")
            else: unique_order_nums.append(str(row['PO']) + "-" + str(row['MODEL']))
        noMatchesCounter = oneMatch = multipleMatches = 0
        for unique_value in unique_order_nums:
            
            cursor = connection.cursor()
            split_value = str(unique_value).split("-")

            numberOfLineExcel = 0
            filtered_data_sheet = []
            for row in data_sheet_rows:
                if split_value[0] in row['PO'] and split_value[1] == str(row['MODEL']): 
                    filtered_data_sheet.append(row)
                    numberOfLineExcel = numberOfLineExcel + 1
            
            noMatchFlag = False
            for row in filtered_data_sheet: 
                if row['PROMISE_DATE'] != filtered_data_sheet[0]['PROMISE_DATE']:
                    utils.print_modified(f'Not all dates are the same for: {split_value}')
                    noMatchFlag = True

            sql = f"SELECT ROUND((FBUY_UNIT_PRICE * (1 - DISCOUNT/100)), 2) as discounted_price, PCFT.CF$_SCHEDULED_DATE, POLT.DEMAND_ORDER_NO, POLT.DEMAND_RELEASE, POLT.* FROM PURCHASE_ORDER_LINE_TAB POLT inner join PURCHASE_ORDER_LINE_PART_CFT PCFT ON POLT.ROWKEY = PCFT.ROWKEY WHERE ROWSTATE != 'Closed' AND ORDER_NO = '{split_value[0]}' AND DESCRIPTION = '{split_value[1]}' ORDER BY PCFT.CF$_SCHEDULED_DATE"

            cursor.execute(sql)
            result = cursor.fetchall()

            if(len(result) > 0):
                for row in result:
                    if row[1] != result[0][1]:
                        utils.print_modified(f"Not all scheduled dates in IFS are the same for: {split_value}")
                        noMatchFlag = True

            if(noMatchFlag): continue

            if(len(result) == 0): 
                utils.print_modified(f"No Results Found for Order No: {split_value[0]} and Total: {split_value[1]}")
                noMatchesCounter = noMatchesCounter + numberOfLineExcel
                noMatchFlag = True
                noMatches.append([split_value[0], split_value[1], 'Could not find any results is IFS.'])

            elif(len(result) > 1): 
                ifsDate = "Empty" if result[0][1] == None else result[0][1].strftime('%d/%m/%Y')
                utils.print_modified(f"More Than Result Found for Order No: {split_value[0]} and Total: {split_value[1]}")
                multipleMatches = multipleMatches + numberOfLineExcel
                POMatchesSheet.append([split_value[0], split_value[1], 'Match found with multiple lines in IFS', ifsDate, filtered_data_sheet[0]['PROMISE_DATE'].strftime('%d/%m/%Y'), len(result), len(filtered_data_sheet)])
                for row in result:
                    cursor.execute(f"UPDATE PURCHASE_ORDER_LINE_PART_CFT SET CF$_SCHEDULED_DATE = TO_DATE('{filtered_data_sheet[0]['PROMISE_DATE'].strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ROWKEY = '{row[149]}'")

            else:
                ifsDate = "Empty" if result[0][1] == None else result[0][1].strftime('%d/%m/%Y')
                utils.print_modified(f"One Result Found for Order No: {split_value[0]} and Total: {split_value[1]}")
                oneMatch = oneMatch + numberOfLineExcel
                POMatchesSheet.append([split_value[0], split_value[1], 'Match found with one line in IFS', ifsDate, filtered_data_sheet[0]['PROMISE_DATE'].strftime('%d/%m/%Y'), len(result), len(filtered_data_sheet)])
                cursor.execute(f"UPDATE PURCHASE_ORDER_LINE_PART_CFT SET CF$_SCHEDULED_DATE = TO_DATE('{filtered_data_sheet[0]['PROMISE_DATE'].strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ROWKEY = '{result[0][149]}'")



            if(len(result) != 0 and result[0][2] != None and result[0][3] != None):
                sql = f"SELECT COLT.PROMISED_DELIVERY_DATE, COLT.* FROM CUSTOMER_ORDER_LINE_TAB COLT WHERE COLT.ORDER_NO = '{result[0][2]}' and COLT.REL_NO = {result[0][3]}"
                
                cursor.execute(sql)
                result_CO = cursor.fetchall()

                if(len(result_CO) == 1):
                    utils.print_modified("CO present")
                    new_date = pd.to_datetime(filtered_data_sheet[0]['PROMISE_DATE'], format="%d/%m/%Y") + timedelta(days=14)
                    COMatchesSheet.append([result_CO[0][1], split_value[0], result_CO[0][2], result_CO[0][3], split_value[1], 'Match found with one CO line in IFS', result_CO[0][0].strftime('%d/%m/%Y'), new_date.strftime('%d/%m/%Y'), len(result_CO)])
                    cursor.execute(f"UPDATE CUSTOMER_ORDER_LINE_TAB SET PROMISED_DELIVERY_DATE = TO_DATE('{new_date.strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ORDER_NO = '{result[0][2]}' and REL_NO = {int(result[0][3])}")
                elif(len(result_CO) > 1):
                    utils.print_modified("multiple COs")
                    new_date = pd.to_datetime(filtered_data_sheet[0]['PROMISE_DATE'], format="%d/%m/%Y") + timedelta(days=14)
                    COMatchesSheet.append([result_CO[0][1], split_value[0], result_CO[0][2], result_CO[0][3], split_value[1], 'Match found with multiple CO lines in IFS', result_CO[0][0].strftime('%d/%m/%Y'), new_date.strftime('%d/%m/%Y'), len(result_CO)])
                    cursor.execute(f"UPDATE CUSTOMER_ORDER_LINE_TAB SET PROMISED_DELIVERY_DATE = TO_DATE('{new_date.strftime('%d/%m/%Y')}', 'DD/MM/YYYY') WHERE ORDER_NO = '{result[0][2]}' and REL_NO = {int(result[0][3])}")
                else:
                    utils.print_modified("CO not present")


        utils.print_modified("One match count: " + str(oneMatch))
        utils.print_modified("Mutliple matches count: " + str(multipleMatches))
        utils.print_modified("No match count: " + str(noMatchesCounter))
        utils.print_modified("Number of unique matches: " + str(len(unique_order_nums)))

        if (not debug):
            connection.commit()    

        connection.close()
        utils.update_json_value(config_file_path, ["PARTNER", "current_email_timestamp"], current_email_timestamp)

        timestamp = datetime.now().strftime('%Y-%m-%d_%Hh-%Mm-%Ss')
        excel_file_name = f"results_laurentide_partner_{timestamp}.xlsx"
        excel_file_path = os.path.join(temp_dir, excel_file_name)


        utils.adjust_column_width(POMatchesSheet, workbook, excel_file_path)
        utils.adjust_column_width(COMatchesSheet, workbook, excel_file_path)
        utils.adjust_column_width(noMatches, workbook, excel_file_path)

        workbook.save(excel_file_path)
        workbook.close()

        utils.send_outlook_email_graph(token=outlook_token, attachement_name=excel_file_name, attachment_directory=temp_dir, mailbox_user_principal_name=SENDER_EMAIL,
                                       recipients=recipients, title="Larentide Impact Partner Automation Results", body_text="See attached the excel file indicating the changes in IFS.")
    
    except Exception as e:
        traceback.print_exc()
        utils.print_modified(str(e))
        env_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
        load_dotenv(env_file_path)
        utils.send_outlook_email_graph(token=outlook_token, mailbox_user_principal_name=os.getenv("SENDER_EMAIL"), recipients=os.getenv("ADMIN_EMAIL"), title="Error! Larentide Impact Partner Automation", body_text=f"Error:\n\n{str(e)}\n\n{traceback.format_exc()}")
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            utils.print_modified("Temp dir deleted")
        utils.print_modified("Script completed")

if __name__ == "__main__":
    import utils as utils
    main(None, utils)